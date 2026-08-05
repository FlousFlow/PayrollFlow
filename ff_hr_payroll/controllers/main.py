# -*- coding: utf-8 -*-
import io

from odoo import http
from odoo.http import request

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class HrPayrollExportController(http.Controller):

    @http.route('/ff_hr_payroll/payslip_register_xlsx', type='http',
                auth='user', methods=['GET'], website=False)
    def payslip_register_xlsx(self, ids=None, **kw):
        """Download the aggregate payroll register as an Excel file."""
        if openpyxl is None:
            return request.not_found()
        if not ids:
            return request.not_found()
        payslips = request.env['hr.payslip'].browse(
            [int(x) for x in ids.split(',') if x.strip().isdigit()])
        payslips = payslips.exists().filtered(
            lambda p: p.company_id in request.env.companies)
        if not payslips:
            return request.not_found()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Payslip Register'
        bold = Font(bold=True)
        header_fill = PatternFill('solid', fgColor='DDEBF7')
        headers = ['Employee', 'Gross', 'Deductions', 'Net',
                   'From', 'To', 'Status', 'Reference']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = bold
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center')
        row = 2
        for p in payslips:
            ws.cell(row=row, column=1, value=p.employee_id.name or '')
            ws.cell(row=row, column=2, value=p.gross_wage)
            ws.cell(row=row, column=3, value=p.total_deductions)
            ws.cell(row=row, column=4, value=p.net_wage)
            ws.cell(row=row, column=5,
                    value=p.date_from.strftime('%Y-%m-%d') if p.date_from else '')
            ws.cell(row=row, column=6,
                    value=p.date_to.strftime('%Y-%m-%d') if p.date_to else '')
            ws.cell(row=row, column=7, value=p.state)
            ws.cell(row=row, column=8, value=p.number or p.name or '')
            row += 1
        # totals
        ws.cell(row=row, column=2, value=sum(p.gross_wage for p in payslips)).font = bold
        ws.cell(row=row, column=3, value=sum(p.total_deductions for p in payslips)).font = bold
        ws.cell(row=row, column=4, value=sum(p.net_wage for p in payslips)).font = bold
        ws.cell(row=row, column=1, value='Total').font = bold
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return request.make_response(
            stream.read(),
            headers=[
                ('Content-Type',
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition',
                 'attachment; filename="payslip_register.xlsx"'),
            ])
